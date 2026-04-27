"""Benchmark clustering analysis for STEB.

Discovers which datasets within a task type measure similar constructs
(i.e., rank models the same way) using pairwise Spearman rank correlations
and hierarchical clustering, following the methodology from OLMo 3 Section 3.3.1.
"""
import argparse
import json
import os
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import yaml

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from scipy.cluster.hierarchy import dendrogram, fcluster, leaves_list, linkage
from scipy.spatial.distance import squareform
from scipy.stats import spearmanr

# Add project root so we can import steb
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from steb.core import get_supported_datasets
from steb.utils import RESULTS_DIR

# ============================================================
# Configuration
# ============================================================

# Maps task name -> primary metric
TASK_METRICS: Dict[str, str] = {
    "clustering": "v_measure",
    "all_to_all_pair_classification": "auc",
    "pre_defined_pair_classification": "auc",
    "order_alignment": "distractor_acc_mean",
    "retrieval": "mrr",
    "probing": "average",
}

# Recognised --oa_variant values for cluster YAML entries.
# Each maps the variant name to the metric used for the order_alignment task
# when that variant is selected. --oa_variant only controls the metric; use
# --oa_only to additionally restrict an entry to the order_alignment task.
OA_VARIANT_METRICS: Dict[str, str] = {
    "distractor": "distractor_acc_mean",
    "acc": "acc_mean",
}


@dataclass(frozen=True)
class ClusterEntry:
    """A parsed cluster YAML entry.

    Attributes:
        name: The dataset name.
        oa_variant: Which order_alignment metric to use for this entry —
            None (use TASK_METRICS default), "distractor", or "acc".
        oa_only: If True, only the order_alignment task contributes to the
            cluster table for this entry; other tasks the dataset declares
            are dropped.
    """
    name: str
    oa_variant: Optional[str] = None
    oa_only: bool = False

# Datasets where the label is primarily semantic (topic, sentiment, content)
# rather than stylistic. Excluded from analysis by default.
SEMANTIC_DATASETS = {
    # Topic / content-based
    "20_Newsgroups_Fixed",
    "ag_news",
    "reuters21578",
    # Sentiment / emotion
    "emotion",
    "financial_phrasebank",
    "twitter-airline-sentiment",
    "yelp_polarity",
    # MISC (not semantic, but I don't want it)
    "probing_blog_small",
    "dummy_retrieval",
    "dummy_order_alignment",
    "fast_baseline_compare_234grams",
    "lftk_sweep",
    "lftk_sweep_fast",
    "lftk_sweep_fast_surfaceavg",
    "lftk_sweep_fast_surfacepos",
}

# Non-English datasets. Excluded from analysis by default.
NON_ENGLISH_DATASETS = {
    # PAN13
    "pan13_authorship_verification_greek_test",
    "pan13_authorship_verification_spanish_test",
    # PAN14
    "pan14_authorship_verification_corpus1_dutch_essays_test",
    "pan14_authorship_verification_corpus1_dutch_reviews_test",
    "pan14_authorship_verification_corpus1_greek_articles_test",
    "pan14_authorship_verification_corpus1_spanish_articles_test",
    "pan14_authorship_verification_corpus2_dutch_essays_test",
    "pan14_authorship_verification_corpus2_dutch_reviews_test",
    "pan14_authorship_verification_corpus2_greek_articles_test",
    "pan14_authorship_verification_corpus2_spanish_articles_test",
    # PAN15
    "pan15_authorship_verification_dutch_test",
    "pan15_authorship_verification_greek_test",
    "pan15_authorship_verification_spanish_test",
    # PAN18
    "pan18_cross_domain_authorship_attribution_french",
    "pan18_cross_domain_authorship_attribution_italian",
    "pan18_cross_domain_authorship_attribution_polish",
    "pan18_cross_domain_authorship_attribution_spanish",
}

EXCLUDED_DATASETS = SEMANTIC_DATASETS | NON_ENGLISH_DATASETS

# Models to exclude from analysis (e.g. broken runs, non-comparable baselines).
EXCLUDED_MODELS: set[str] = set()
EXCLUDED_MODELS.add("avgs_typetoken_read.yaml")
# EXCLUDED_MODELS.add("surface_pos.yaml")
EXCLUDED_MODELS.add("lftk")
EXCLUDED_MODELS.add("tfidf")
EXCLUDED_MODELS.add("tfidfngrams")

LOW_CONFIDENCE_THRESHOLD = 10


# ============================================================
# Shared utilities
# ============================================================


def _warn_missing_metric(
    dataset: str,
    task: str,
    metric: str,
    seen: set,
) -> None:
    """Print a deduped warning to stderr when a metric is missing for a run.

    Each (dataset, task, metric) combination is warned about at most once
    per call site (deduped via the caller-provided ``seen`` set), so model
    counts are not part of the dedupe key — a single warning per data
    triple is enough to alert the user.
    """
    key = (dataset, task, metric)
    if key in seen:
        return
    seen.add(key)
    print(
        f"  WARNING: ignoring runs for dataset '{dataset}' / task '{task}' "
        f"that are missing the '{metric}' metric.",
        file=sys.stderr,
    )


def _resolve_metric_for_entry(
    task: str,
    entry: ClusterEntry,
) -> Optional[str]:
    """Pick which metric an entry contributes to a (task, metric) column.

    --oa_variant only matters for the order_alignment task; for all other
    tasks the default TASK_METRICS metric is used. Returns None when the
    task is unknown to TASK_METRICS (so the caller should skip).
    """
    if task == "order_alignment" and entry.oa_variant is not None:
        return OA_VARIANT_METRICS[entry.oa_variant]
    return TASK_METRICS.get(task)


def discover_all_scores(
    results_dir: str,
    include_excluded: bool = False,
) -> Dict[Tuple[str, str, str], Dict[str, Dict[str, float]]]:
    """Scan the results directory and collect top-level metrics across tasks.

    Respects EXCLUDED_DATASETS, EXCLUDED_MODELS, and NON_ENGLISH_DATASETS
    filtering. Collects every (dataset, task, episode_config, model)
    combination found, returning the top-level scalar metrics from each
    run's metrics.json so callers can pick whichever metric they need
    (e.g. acc_mean vs distractor_acc_mean for order_alignment). Nested
    values (e.g. _per_label, submetrics) are dropped to keep memory low.

    This is the only function that walks the filesystem; other consumers
    (e.g. ``discover_scores`` for the auto-cluster path) build on top of
    it.

    Args:
        results_dir: Path to the root results directory.
        include_excluded: If True, include semantic and non-English datasets.

    Returns:
        A dict mapping (dataset, task, episode_config) to a dict mapping
        model name to that run's flat (scalar-only) metrics dict.
    """
    results_path = Path(results_dir)
    if not results_path.exists():
        return {}

    # Build a map of dataset -> set of tasks it supports
    dataset_tasks: Dict[str, set] = {}
    for task_name in TASK_METRICS:
        for ds in get_supported_datasets(task_name):
            dataset_tasks.setdefault(ds, set()).add(task_name)

    # Collect: (dataset, task, episode_config) -> {model: metrics_dict}
    rows: Dict[Tuple[str, str, str], Dict[str, Dict[str, Any]]] = {}

    for dataset_dir in sorted(results_path.iterdir()):
        if not dataset_dir.is_dir():
            continue

        dataset_name = dataset_dir.name
        if dataset_name not in dataset_tasks:
            continue
        if not include_excluded and dataset_name in EXCLUDED_DATASETS:
            continue

        for model_dir in sorted(dataset_dir.iterdir()):
            if not model_dir.is_dir():
                continue
            if model_dir.name in EXCLUDED_MODELS:
                continue

            for ep_dir in sorted(model_dir.iterdir()):
                if not ep_dir.is_dir():
                    continue

                for task_name in dataset_tasks[dataset_name]:
                    metrics_file = ep_dir / task_name / "metrics.json"
                    if not metrics_file.exists():
                        continue

                    with open(metrics_file) as f:
                        metrics = json.load(f)

                    # Strip nested values (e.g. _per_label, submetrics) — only
                    # top-level scalar metrics are needed by current consumers.
                    top_level_metrics = {
                        k: v for k, v in metrics.items() if not isinstance(v, dict)
                    }

                    key = (dataset_name, task_name, ep_dir.name)
                    rows.setdefault(key, {})[model_dir.name] = top_level_metrics

    return rows


def discover_scores(
    results_dir: str,
    task_name: str,
    primary_metric: str,
    episode_params: Optional[str],
    include_excluded: bool = False,
) -> pd.DataFrame:
    """Build a models × datasets DataFrame for one task and one metric.

    Thin wrapper around ``discover_all_scores`` that filters to the given
    task, picks the first matching episode_params per (dataset, model),
    and pivots into a DataFrame.

    Args:
        results_dir: Path to the root results directory.
        task_name: The CLI task name (e.g. 'clustering').
        primary_metric: The metric to extract from each metrics dict.
        episode_params: Episode params filter like '1_50'. If None, picks
            the first episode params found per dataset-model pair (in
            sorted order).
        include_excluded: If True, include semantic and non-English datasets.

    Returns:
        A DataFrame with models as rows and datasets as columns.
    """
    all_scores = discover_all_scores(results_dir, include_excluded)

    scores: Dict[str, Dict[str, float]] = {}
    # Iterate sorted so the "first ep_config wins per (model, dataset)"
    # rule mirrors the previous sorted-iterdir traversal.
    for (dataset, task, ep_config), model_metrics in sorted(all_scores.items()):
        if task != task_name:
            continue
        if episode_params and ep_config != episode_params:
            continue
        for model, metrics in model_metrics.items():
            if dataset in scores.get(model, {}):
                continue  # earlier ep_config already filled this (model, dataset)
            value = metrics.get(primary_metric)
            if value is not None:
                scores.setdefault(model, {})[dataset] = value

    if not scores:
        return pd.DataFrame()

    return pd.DataFrame(scores).T.rename_axis("model")


# ============================================================
# Auto-cluster analysis (--task / --all-tasks)
# ============================================================


def compute_correlation_matrix(
    df: pd.DataFrame,
) -> pd.DataFrame:
    """Compute pairwise Spearman rank correlations between dataset columns.

    Args:
        df: Score matrix with models as rows and datasets as columns.

    Returns:
        A symmetric DataFrame of pairwise Spearman correlations.
    """
    n = len(df.columns)
    corr = np.ones((n, n))

    for i in range(n):
        for j in range(i + 1, n):
            rho, _ = spearmanr(df.iloc[:, i], df.iloc[:, j])
            if np.isnan(rho):
                rho = 0.0
            corr[i, j] = corr[j, i] = rho

    return pd.DataFrame(corr, index=df.columns, columns=df.columns)


def plot_dendrogram(
    corr_matrix: pd.DataFrame,
    task_name: str,
    output_path: str,
) -> np.ndarray:
    """Plot and save a dendrogram from the correlation matrix.

    Args:
        corr_matrix: Pairwise Spearman correlation matrix.
        task_name: Task name for the plot title.
        output_path: File path to save the figure.

    Returns:
        The linkage matrix from hierarchical clustering.
    """
    dist = np.clip(1 - corr_matrix.values, 0, 2)
    np.fill_diagonal(dist, 0)
    dist = (dist + dist.T) / 2
    Z = linkage(squareform(dist), method="ward")

    fig, ax = plt.subplots(figsize=(max(10, len(corr_matrix) * 0.8), 6))
    dendrogram(Z, labels=corr_matrix.columns.tolist(), ax=ax, leaf_rotation=90, leaf_font_size=9)
    ax.set_title(f"Dataset Clustering — {task_name}")
    ax.set_ylabel("Distance (1 − Spearman ρ)")
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)

    return Z


def plot_heatmap(
    corr_matrix: pd.DataFrame,
    Z: np.ndarray,
    task_name: str,
    output_path: str,
) -> None:
    """Plot and save a heatmap of the correlation matrix ordered by clustering.

    Args:
        corr_matrix: Pairwise Spearman correlation matrix.
        Z: Linkage matrix from hierarchical clustering.
        task_name: Task name for the plot title.
        output_path: File path to save the figure.
    """
    order = leaves_list(Z)
    ordered_labels = [corr_matrix.columns[i] for i in order]
    ordered_corr = corr_matrix.loc[ordered_labels, ordered_labels]

    n = len(ordered_labels)
    fig, ax = plt.subplots(figsize=(max(8, n * 0.7), max(7, n * 0.6)))
    im = ax.imshow(ordered_corr.values, cmap="RdBu_r", vmin=-1, vmax=1, aspect="auto")
    ax.set_xticks(range(n))
    ax.set_yticks(range(n))
    ax.set_xticklabels(ordered_labels, rotation=90, fontsize=8)
    ax.set_yticklabels(ordered_labels, fontsize=8)
    ax.set_title(f"Spearman Correlation — {task_name}")

    for i in range(n):
        for j in range(n):
            val = ordered_corr.values[i, j]
            color = "white" if abs(val) > 0.7 else "black"
            ax.text(j, i, f"{val:.2f}", ha="center", va="center", fontsize=6, color=color)

    fig.colorbar(im, ax=ax, shrink=0.8)
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def build_summary(
    Z: np.ndarray,
    corr_matrix: pd.DataFrame,
    n_models: int,
    threshold: float = 1.0,
) -> dict:
    """Build a summary of discovered clusters at a given distance threshold.

    Args:
        Z: Linkage matrix from hierarchical clustering.
        corr_matrix: Pairwise Spearman correlation matrix.
        n_models: Number of models with complete results.
        threshold: Distance threshold for forming flat clusters.

    Returns:
        A dictionary with cluster assignments, metadata, and confidence flag.
    """
    labels = fcluster(Z, t=threshold, criterion="distance")
    datasets = corr_matrix.columns.tolist()

    clusters: Dict[str, List[str]] = {}
    for dataset, label in zip(datasets, labels):
        clusters.setdefault(f"cluster_{label}", []).append(dataset)

    return {
        "n_models": n_models,
        "n_datasets": len(datasets),
        "threshold": threshold,
        "n_clusters": len(clusters),
        "clusters": clusters,
        "low_confidence": n_models < LOW_CONFIDENCE_THRESHOLD,
    }


def aggregate_scores(
    df: pd.DataFrame,
    clusters: Dict[str, List[str]],
) -> pd.DataFrame:
    """Compute cluster-aware aggregated scores for each model.

    For each model: macro-average scores within each cluster, then average
    across clusters. This gives equal weight to each cluster regardless of
    how many datasets it contains.

    Args:
        df: Score matrix with models as rows and datasets as columns.
        clusters: Dict mapping cluster names to lists of dataset names.

    Returns:
        A DataFrame with columns: one per cluster (cluster mean), plus
        'task_score' (the final aggregated score). Indexed by model name.
    """
    result = pd.DataFrame(index=df.index)

    for cluster_name, datasets in sorted(clusters.items()):
        present = [d for d in datasets if d in df.columns]
        if present:
            result[cluster_name] = df[present].mean(axis=1)

    result["task_score"] = result.mean(axis=1)
    return result


def analyze_task(
    results_dir: str,
    task_name: str,
    primary_metric: str,
    episode_params: Optional[str],
    output_dir: str,
    min_models: int,
    include_excluded: bool = False,
    threshold: float = 1.0,
    complete_datasets: bool = False,
) -> Optional[pd.Series]:
    """Run the full clustering analysis for a single task.

    Args:
        results_dir: Path to the root results directory.
        task_name: The CLI task name.
        primary_metric: The metric to extract and rank by.
        episode_params: Episode params filter (e.g. '1_50').
        output_dir: Directory to write output files.
        min_models: Minimum number of models with complete results.
        include_excluded: If True, include semantic and non-English datasets.
        threshold: Distance threshold for flat clustering.
        complete_datasets: If True, drop datasets with missing models instead
            of dropping models with missing datasets.

    Returns:
        A Series of per-model task scores, or None if the task was skipped.
    """
    print(f"\n{'='*60}")
    print(f"Task: {task_name}")
    print(f"{'='*60}")

    df = discover_scores(results_dir, task_name, primary_metric, episode_params, include_excluded)
    if df.empty:
        print(f"  No results found. Skipping.")
        return None

    print(f"  Raw matrix: {len(df)} models × {len(df.columns)} datasets")

    raw_df = df
    if complete_datasets:
        df = df.dropna(axis=1, how="any")
        dropped_datasets = set(raw_df.columns) - set(df.columns)
        if dropped_datasets:
            print(f"  Dropped {len(dropped_datasets)} incomplete dataset(s): {sorted(dropped_datasets)}")
        if len(df) < min_models:
            print(f"  Only {len(df)} models (need {min_models}). Skipping.")
            return None
    else:
        df = df.dropna(axis=0, how="any")
        if len(df) < min_models:
            print(f"  Only {len(df)} models with complete results (need {min_models}). Skipping.")
            dropped = raw_df[raw_df.isna().any(axis=1)]
            if not dropped.empty:
                print(f"  Missing evaluations:")
                for model in dropped.index:
                    missing = [d for d in dropped.columns if pd.isna(dropped.at[model, d])]
                    for dataset in missing:
                        print(f"    ({model}, {dataset})")
            return None

    print(f"  Complete matrix: {len(df)} models × {len(df.columns)} datasets")

    if len(df) < LOW_CONFIDENCE_THRESHOLD:
        print(f"  Warning: fewer than {LOW_CONFIDENCE_THRESHOLD} models — results are low-confidence.")

    os.makedirs(output_dir, exist_ok=True)

    # Save score matrix
    scores_path = os.path.join(output_dir, f"{task_name}_scores.csv")
    df.to_csv(scores_path)
    print(f"  Saved score matrix: {scores_path}")

    if len(df.columns) < 2:
        print(f"  Only {len(df.columns)} dataset(s). Skipping clustering, using raw scores.")
        task_score = df.iloc[:, 0]
        task_score.name = "task_score"
        agg_path = os.path.join(output_dir, f"{task_name}_aggregated.csv")
        task_score.to_csv(agg_path)
        print(f"  Saved aggregated scores: {agg_path}")
        print(f"  Task scores:")
        for model in task_score.index:
            print(f"    {model}: {task_score.loc[model]:.4f}")
        return task_score

    # Compute correlations
    corr = compute_correlation_matrix(df)
    corr_path = os.path.join(output_dir, f"{task_name}_correlation.csv")
    corr.to_csv(corr_path)
    print(f"  Saved correlation matrix: {corr_path}")

    # Dendrogram
    dendro_path = os.path.join(output_dir, f"{task_name}_dendrogram.png")
    Z = plot_dendrogram(corr, task_name, dendro_path)
    print(f"  Saved dendrogram: {dendro_path}")

    # Heatmap
    heatmap_path = os.path.join(output_dir, f"{task_name}_heatmap.png")
    plot_heatmap(corr, Z, task_name, heatmap_path)
    print(f"  Saved heatmap: {heatmap_path}")

    # Summary
    summary = build_summary(Z, corr, n_models=len(df), threshold=threshold)
    summary_path = os.path.join(output_dir, f"{task_name}_summary.json")
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"  Saved summary: {summary_path}")
    print(f"  Clusters (threshold={summary['threshold']}):")
    for name, members in summary["clusters"].items():
        print(f"    {name}: {members}")

    # Aggregated scores
    agg = aggregate_scores(df, summary["clusters"])
    agg_path = os.path.join(output_dir, f"{task_name}_aggregated.csv")
    agg.to_csv(agg_path)
    print(f"  Saved aggregated scores: {agg_path}")
    print(f"  Task scores (cluster-aware):")
    for model in agg.index:
        print(f"    {model}: {agg.loc[model, 'task_score']:.4f}")

    return agg["task_score"]


def print_summary_table(
    task_scores: Dict[str, pd.Series],
    task_metrics: Dict[str, str],
    output_dir: str,
) -> None:
    """Print a Markdown table summarizing per-model scores across tasks.

    Bolds the best score in each column. Saves the table to a text file.

    Args:
        task_scores: Mapping from task name to a Series of per-model task scores.
        task_metrics: Mapping from task name to metric name (for column headers).
        output_dir: Directory to save the summary table file.
    """
    if not task_scores:
        return

    columns = {
        f"{task} ({task_metrics[task]})": scores
        for task, scores in task_scores.items()
    }
    df = pd.DataFrame(columns)
    df.index.name = "Model"

    # Bold the best value in each column
    for col in df.columns:
        valid = df[col].dropna()
        if valid.empty:
            df[col] = "—"
            continue
        best_idx = valid.idxmax()
        df[col] = df[col].apply(lambda x: f"{x:.4f}" if pd.notna(x) else "—")
        df.at[best_idx, col] = f"**{df.at[best_idx, col]}**"

    table = df.to_markdown()

    print(f"\n{'='*60}")
    print("Summary")
    print(f"{'='*60}\n")
    print(table)
    print()

    os.makedirs(output_dir, exist_ok=True)
    table_path = os.path.join(output_dir, "summary_table.md")
    with open(table_path, "w") as f:
        f.write(table + "\n")
    print(f"Saved summary table: {table_path}")


# ============================================================
# Manual cluster YAML parsing
# ============================================================


def parse_cluster_entry(
    entry: str,
) -> ClusterEntry:
    """Parse a cluster YAML dataset entry into a ClusterEntry.

    Supports an optional --oa_variant VALUE flag and an optional
    --oa_only flag, in any order. --oa_variant selects which
    order_alignment metric to use for this entry (one of
    OA_VARIANT_METRICS); --oa_only restricts the entry to just the
    order_alignment task. Either may appear without the other.

    Args:
        entry: A raw entry string from the cluster YAML.

    Returns:
        A ClusterEntry capturing the dataset name and any flags.
    """
    name, *tokens = entry.split()
    oa_variant: Optional[str] = None
    oa_only = False
    it = iter(tokens)
    for tok in it:
        if tok == "--oa_only":
            oa_only = True
        elif tok == "--oa_variant":
            oa_variant = next(it, None)
            if oa_variant not in OA_VARIANT_METRICS:
                raise ValueError(
                    f"Invalid cluster entry '{entry}': --oa_variant expected "
                    f"one of {sorted(OA_VARIANT_METRICS)}, got {oa_variant!r}."
                )
        else:
            raise ValueError(
                f"Invalid cluster entry '{entry}': unexpected token {tok!r}."
            )
    return ClusterEntry(name=name, oa_variant=oa_variant, oa_only=oa_only)


def load_manual_clusters(
    clusters_path: str,
) -> Dict[str, List[ClusterEntry]]:
    """Load manual dataset clusters from a YAML file.

    Args:
        clusters_path: Path to the YAML file with cluster definitions.

    Returns:
        A dict mapping cluster name to a list of ClusterEntry records.
    """
    with open(clusters_path) as f:
        raw = yaml.safe_load(f)

    clusters: Dict[str, List[ClusterEntry]] = {}
    for name, config in raw.items():
        clusters[name] = [parse_cluster_entry(entry) for entry in config["datasets"]]
    return clusters


# ============================================================
# Manual cluster table building (--manual-clusters)
# ============================================================


def _collect_cluster_table_scores(
    all_scores: Dict[Tuple[str, str, str], Dict[str, Dict[str, float]]],
    dataset_to_entries: Dict[str, List[Tuple[str, ClusterEntry]]],
    episode_params: Optional[str],
) -> Tuple[
    Dict[Tuple[str, str, str, str], Dict[str, float]],
    set,
]:
    """Walk discovered runs and bucket scores by (cluster, task, metric, dataset).

    Honours each entry's --oa_only and --oa_variant flags. Emits a deduped
    stderr warning whenever a run is dropped because its metrics.json
    lacks the resolved metric. Returns the per-dataset score buckets plus
    the set of (cluster, dataset) pairs where an --oa_only entry actually
    contributed data.
    """
    per_dataset: Dict[Tuple[str, str, str, str], Dict[str, float]] = {}
    warned_missing: set = set()
    oa_only_contributed: set = set()

    for (dataset, task, ep_config), model_metrics in all_scores.items():
        if episode_params and ep_config != episode_params:
            continue
        if dataset not in dataset_to_entries:
            continue

        for cluster_name, entry in dataset_to_entries[dataset]:
            if entry.oa_only and task != "order_alignment":
                continue

            metric_key = _resolve_metric_for_entry(task, entry)
            if metric_key is None:
                continue

            scores: Dict[str, float] = {}
            for model, metrics in model_metrics.items():
                value = metrics.get(metric_key)
                if value is None:
                    _warn_missing_metric(dataset, task, metric_key, warned_missing)
                    continue
                scores[model] = value

            if not scores:
                continue

            if entry.oa_only:
                oa_only_contributed.add((cluster_name, dataset))

            key = (cluster_name, task, metric_key, dataset)
            per_dataset.setdefault(key, {}).update(scores)

    return per_dataset, oa_only_contributed


def _filter_incomplete_datasets(
    cluster_col_groups: Dict[Tuple[str, str, str], Dict[str, Dict[str, float]]],
) -> None:
    """Mutate cluster_col_groups in place, dropping datasets that don't have all models."""
    for group_key, dataset_scores in cluster_col_groups.items():
        all_models: set[str] = set()
        for model_scores in dataset_scores.values():
            all_models.update(model_scores.keys())

        complete = {
            ds: scores for ds, scores in dataset_scores.items()
            if set(scores.keys()) == all_models
        }
        dropped = set(dataset_scores.keys()) - set(complete.keys())
        if dropped:
            cluster_name, task, metric_key = group_key
            print(f"  Manual cluster '{cluster_name}' / {task} ({metric_key}): "
                  f"dropped {len(dropped)} incomplete dataset(s): {sorted(dropped)}")
        cluster_col_groups[group_key] = complete


def _dataframes_from_groups(
    cluster_col_groups: Dict[Tuple[str, str, str], Dict[str, Dict[str, float]]],
) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Dict[str, List[str]]]]:
    """Turn the (cluster, task, metric) groups into one DataFrame per cluster.

    Returns the per-cluster tables plus the per-cluster mapping from
    column header to the list of contributing datasets.
    """
    cluster_model_col: Dict[str, Dict[str, Dict[Tuple[str, str], List[float]]]] = {}
    cluster_col_datasets: Dict[str, Dict[Tuple[str, str], set]] = {}

    for (cluster_name, task, metric_key), dataset_scores in cluster_col_groups.items():
        col_id = (task, metric_key)
        for dataset, model_scores in dataset_scores.items():
            cluster_col_datasets.setdefault(cluster_name, {}).setdefault(col_id, set()).add(dataset)
            for model, score in model_scores.items():
                cluster_model_col.setdefault(cluster_name, {})
                cluster_model_col[cluster_name].setdefault(model, {})
                cluster_model_col[cluster_name][model].setdefault(col_id, []).append(score)

    tables: Dict[str, pd.DataFrame] = {}
    column_datasets: Dict[str, Dict[str, List[str]]] = {}
    for cluster_name, model_data in sorted(cluster_model_col.items()):
        records = {}
        for model, col_scores in sorted(model_data.items()):
            records[model] = {
                f"{task} ({metric_key})": np.mean(scores)
                for (task, metric_key), scores in sorted(col_scores.items())
            }
        df = pd.DataFrame(records).T
        df.index.name = "model"
        tables[cluster_name] = df

        col_ds = {}
        for col_id in sorted(cluster_col_datasets.get(cluster_name, {})):
            task, metric_key = col_id
            col_ds[f"{task} ({metric_key})"] = sorted(cluster_col_datasets[cluster_name][col_id])
        column_datasets[cluster_name] = col_ds

    return tables, column_datasets


def build_manual_cluster_tables(
    results_dir: str,
    clusters: Dict[str, List[ClusterEntry]],
    episode_params: Optional[str],
    include_excluded: bool = False,
    complete_datasets: bool = False,
) -> tuple[Dict[str, pd.DataFrame], Dict[str, Dict[str, List[str]]]]:
    """Build one table per manual cluster.

    For each cluster, produces a DataFrame where rows are models and columns
    are "task (metric)" — since order_alignment can appear under different
    metrics depending on each entry's --oa_variant, the metric is part of the
    column key. Each cell is the average metric across datasets in that
    cluster that contribute to the column.

    Per-entry flags decouple two concerns:
      - --oa_variant only changes which order_alignment metric is used
        (via OA_VARIANT_METRICS); the entry still contributes to all
        tasks the dataset declares.
      - --oa_only restricts an entry to the order_alignment task; other
        tasks the dataset declares are dropped for that entry.

    Args:
        results_dir: Path to the root results directory.
        clusters: Mapping from cluster name to a list of ClusterEntry
            records.
        episode_params: Episode params filter (e.g. '1_50').
        include_excluded: If True, include semantic and non-English datasets.
        complete_datasets: If True, within each (cluster, column) group, drop
            datasets that not all models have results for.

    Returns:
        A tuple of:
          - A dict mapping cluster name to a DataFrame whose rows are
            models and whose columns are keyed by (task, metric).
          - A dict mapping cluster name to a dict of column name to
            list of dataset names included in that column.
    """
    all_scores = discover_all_scores(results_dir, include_excluded)
    if not all_scores:
        return {}, {}

    # Invert clusters: dataset -> list of (cluster_name, entry). A dataset
    # may legitimately appear in more than one cluster (e.g. once plain in
    # 'style', once with --oa_only in 'style_vs_content').
    dataset_to_entries: Dict[str, List[Tuple[str, ClusterEntry]]] = {}
    for cluster_name, entries in clusters.items():
        for entry in entries:
            dataset_to_entries.setdefault(entry.name, []).append((cluster_name, entry))

    per_dataset, oa_only_contributed = _collect_cluster_table_scores(
        all_scores, dataset_to_entries, episode_params,
    )

    # Warn for --oa_only entries that produced no order_alignment data
    # (e.g. dataset has no results dir, or doesn't declare order_alignment).
    for cluster_name, entries in clusters.items():
        for entry in entries:
            if entry.oa_only and (cluster_name, entry.name) not in oa_only_contributed:
                print(
                    f"  WARNING: cluster '{cluster_name}' entry "
                    f"'{entry.name} --oa_only' produced no order_alignment results.",
                    file=sys.stderr,
                )

    # Group by (cluster, task, metric) to find all models and datasets per column.
    cluster_col_groups: Dict[Tuple[str, str, str], Dict[str, Dict[str, float]]] = {}
    for (cluster_name, task, metric_key, dataset), model_scores in per_dataset.items():
        group_key = (cluster_name, task, metric_key)
        cluster_col_groups.setdefault(group_key, {})
        cluster_col_groups[group_key][dataset] = model_scores

    if complete_datasets:
        _filter_incomplete_datasets(cluster_col_groups)

    return _dataframes_from_groups(cluster_col_groups)


def print_manual_cluster_tables(
    tables: Dict[str, pd.DataFrame],
    column_datasets: Dict[str, Dict[str, List[str]]],
    output_dir: str,
) -> None:
    """Print manual cluster tables to stdout and save as markdown.

    Args:
        tables: Mapping from cluster name to DataFrame whose rows are
            models and whose columns are keyed by (task, metric).
        column_datasets: Mapping from cluster name to dict of column name
            to list of dataset names included in that column.
        output_dir: Directory to save markdown files.
    """
    os.makedirs(output_dir, exist_ok=True)

    for cluster_name, df in tables.items():
        print(f"\n{'='*60}")
        print(f"Manual Cluster — {cluster_name}")
        print(f"{'='*60}\n")

        col_ds = column_datasets.get(cluster_name, {})
        for col_name, datasets in col_ds.items():
            print(f"  {col_name}: {', '.join(datasets)}")
        print()

        print(df.to_markdown())
        print()

        table_path = os.path.join(output_dir, f"manual_clusters_{cluster_name}.md")
        with open(table_path, "w") as f:
            f.write(df.to_markdown() + "\n")
        print(f"Saved: {table_path}")


# ============================================================
# Excel export (--export-excel)
# ============================================================


def _highlight_best_two(
    ws,
    cells: List[Tuple[int, int]],
    bold_font,
    underline_font,
) -> None:
    """Bold the cell with the largest numeric value, underline the second.

    Args:
        ws: An openpyxl worksheet.
        cells: List of (row, col) coordinates to consider. Non-numeric
            cells are ignored. No-op if fewer than 2 numeric values.
        bold_font, underline_font: Pre-built openpyxl Font objects.
    """
    vals: List[Tuple[float, int, int]] = []
    for row, col in cells:
        v = ws.cell(row=row, column=col).value
        if isinstance(v, (int, float)):
            vals.append((v, row, col))
    if len(vals) < 2:
        return
    vals.sort(key=lambda x: x[0], reverse=True)
    ws.cell(row=vals[0][1], column=vals[0][2]).font = bold_font
    ws.cell(row=vals[1][1], column=vals[1][2]).font = underline_font


def _build_scores_records(
    rows: Dict[Tuple[str, str, str], Dict[str, Dict[str, float]]],
) -> List[Dict[str, object]]:
    """Flatten discovered runs into per-row records for the 'scores' sheet.

    One record per (dataset, task, episode_config); each model's primary
    metric is added as a column. Emits a deduped warning when a run is
    missing the expected metric.
    """
    records: List[Dict[str, object]] = []
    warned_missing: set = set()
    for (dataset, task, episode_config), model_metrics in sorted(rows.items()):
        primary_metric = TASK_METRICS[task]
        record: Dict[str, object] = {
            "dataset": dataset,
            "task": task,
            "episode_config": episode_config,
            "primary_metric": primary_metric,
        }
        for model, metrics in model_metrics.items():
            value = metrics.get(primary_metric)
            if value is None:
                _warn_missing_metric(dataset, task, primary_metric, warned_missing)
                continue
            record[model] = value
        records.append(record)
    return records


def _write_scores_sheet(
    writer,
    scores_df: pd.DataFrame,
    n_meta_cols: int,
    n_model_cols: int,
    bold_font,
    underline_font,
) -> None:
    """Write the 'scores' sheet and bold/underline the best two per row."""
    scores_df.to_excel(writer, sheet_name="scores", index=False)
    ws = writer.sheets["scores"]
    model_col_start = n_meta_cols + 1  # 1-indexed, after meta columns
    model_col_end = model_col_start + n_model_cols - 1
    for row_idx in range(2, len(scores_df) + 2):  # skip header
        cells = [(row_idx, c) for c in range(model_col_start, model_col_end + 1)]
        _highlight_best_two(ws, cells, bold_font, underline_font)


def _write_summary_sheet(
    writer,
    task_scores: Dict[str, pd.Series],
    task_metrics: Dict[str, str],
    bold_font,
    underline_font,
) -> None:
    """Write the 'summary' sheet (one row per model, one col per task) and
    bold/underline the best two cells per column."""
    columns = {
        f"{task} ({task_metrics[task]})": scores
        for task, scores in task_scores.items()
    }
    summary_df = pd.DataFrame(columns)
    summary_df.index.name = "model"
    summary_df.to_excel(writer, sheet_name="summary")

    ws = writer.sheets["summary"]
    for col_idx in range(2, len(summary_df.columns) + 2):  # skip index col
        cells = [(r, col_idx) for r in range(2, len(summary_df) + 2)]
        _highlight_best_two(ws, cells, bold_font, underline_font)


def _write_manual_cluster_sheet(
    writer,
    cluster_name: str,
    mc_df: pd.DataFrame,
    col_ds: Dict[str, List[str]],
    bold_font,
    underline_font,
    italic_font,
) -> None:
    """Write one mc_{cluster} sheet: dataset list above, then the table,
    with the best two cells per column bolded/underlined."""
    sheet_name = f"mc_{cluster_name}"[:31]  # Excel 31-char limit

    # Leave rows above the data for the per-column dataset list, plus a blank.
    max_datasets = max((len(ds) for ds in col_ds.values()), default=0)
    data_start_row = max_datasets + 2 if max_datasets > 0 else 0
    mc_df.to_excel(writer, sheet_name=sheet_name, startrow=data_start_row)

    ws = writer.sheets[sheet_name]

    if col_ds:
        ws.cell(row=1, column=1, value="Datasets:").font = italic_font
        for col_idx, col_name in enumerate(mc_df.columns, start=2):
            for ds_idx, ds_name in enumerate(col_ds.get(col_name, [])):
                cell = ws.cell(row=1 + ds_idx, column=col_idx, value=ds_name)
                cell.font = italic_font

    header_row = data_start_row + 1
    for col_idx in range(2, len(mc_df.columns) + 2):
        cells = [(r, col_idx) for r in range(header_row + 1, header_row + len(mc_df) + 1)]
        _highlight_best_two(ws, cells, bold_font, underline_font)


def _autosize_workbook_columns(workbook) -> None:
    """Set every column's width to fit its widest cell value."""
    for ws in workbook.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 2


def export_excel(
    results_dir: str,
    output_path: str,
    include_excluded: bool = False,
    task_scores: Optional[Dict[str, pd.Series]] = None,
    task_metrics: Optional[Dict[str, str]] = None,
    manual_cluster_tables: Optional[Dict[str, pd.DataFrame]] = None,
    manual_cluster_datasets: Optional[Dict[str, Dict[str, List[str]]]] = None,
) -> None:
    """Export all scores to an Excel file.

    Sheets:
      - "scores": one row per (dataset, task, episode_config) with per-model
        metric values. Best per row is bold, second best is underlined.
      - "summary": cluster-aware aggregated scores per model per task
        (if --task or --all-tasks was used).
      - One sheet per cluster named "mc_{cluster}" with manual cluster
        averages (if --manual-clusters was used). Best per column is bold,
        second best is underlined. Includes dataset list per column.

    Args:
        results_dir: Path to the root results directory.
        output_path: Path for the output .xlsx file.
        include_excluded: If True, include semantic and non-English datasets.
        task_scores: Mapping from task name to per-model aggregated scores.
            If provided, written as the "summary" sheet.
        task_metrics: Mapping from task name to metric name (for column headers).
        manual_cluster_tables: Mapping from cluster name to DataFrame of
            manual cluster averages, rows = models, columns keyed by
            (task, metric).
        manual_cluster_datasets: Mapping from cluster name to dict of column
            name to list of dataset names included in that column.
    """
    rows = discover_all_scores(results_dir, include_excluded)
    if not rows:
        print("No scores found. Nothing to export.")
        return

    records = _build_scores_records(rows)
    scores_df = pd.DataFrame(records)

    # Ensure metadata columns come first, then models sorted alphabetically.
    meta_cols = ["dataset", "task", "episode_config", "primary_metric"]
    model_cols = sorted(c for c in scores_df.columns if c not in meta_cols)
    scores_df = scores_df[meta_cols + model_cols]

    from openpyxl.styles import Font
    bold_font = Font(bold=True)
    underline_font = Font(underline="single")
    italic_font = Font(italic=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_scores_sheet(
            writer, scores_df, len(meta_cols), len(model_cols),
            bold_font, underline_font,
        )

        if task_scores and task_metrics:
            _write_summary_sheet(
                writer, task_scores, task_metrics, bold_font, underline_font,
            )

        if manual_cluster_tables:
            col_ds_all = manual_cluster_datasets or {}
            for cluster_name, mc_df in sorted(manual_cluster_tables.items()):
                _write_manual_cluster_sheet(
                    writer, cluster_name, mc_df,
                    col_ds_all.get(cluster_name, {}),
                    bold_font, underline_font, italic_font,
                )

        _autosize_workbook_columns(writer.book)

    n_sheets = 1
    if task_scores:
        n_sheets += 1
    if manual_cluster_tables:
        n_sheets += len(manual_cluster_tables)
    print(f"Exported {len(scores_df)} rows × {len(model_cols)} models ({n_sheets} sheets) to {output_path}")


# ============================================================
# CLI
# ============================================================


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments.

    Returns:
        Parsed arguments namespace.
    """
    parser = argparse.ArgumentParser(
        description="Benchmark clustering analysis for STEB tasks.",
    )
    parser.add_argument(
        "--results-dir",
        default=RESULTS_DIR,
        help="Path to the results directory (default: %(default)s).",
    )
    parser.add_argument(
        "--task",
        choices=list(TASK_METRICS.keys()),
        help="Task to analyze. If omitted with --all-tasks, analyzes all.",
    )
    parser.add_argument(
        "--all-tasks",
        action="store_true",
        help="Run analysis for all task types.",
    )
    parser.add_argument(
        "--metric",
        help="Override the primary metric (default: task-specific).",
    )
    parser.add_argument(
        "--episode-params",
        help="Filter to a specific episode params string (e.g. '1_50').",
    )
    parser.add_argument(
        "--output-dir",
        default="analysis_output",
        help="Directory for output files (default: %(default)s).",
    )
    parser.add_argument(
        "--min-models",
        type=int,
        default=3,
        help="Minimum models with complete results to run analysis (default: %(default)s).",
    )
    parser.add_argument(
        "--include-excluded",
        action="store_true",
        help="Include semantic and non-English datasets that are excluded by default.",
    )
    parser.add_argument(
        "--threshold",
        type=float,
        default=1.0,
        help="Distance threshold for flat clustering (default: %(default)s). "
             "Lower = more clusters (0.5 ≈ ρ≥0.5), higher = fewer (1.5 ≈ ρ≥-0.5).",
    )
    parser.add_argument(
        "--complete-datasets",
        action="store_true",
        help="Instead of dropping models with missing datasets, drop datasets "
             "that not all models have results for. Keeps all models.",
    )
    parser.add_argument(
        "--export-excel",
        metavar="PATH",
        help="Export all scores to an Excel file. Sheet 1 has per-dataset "
             "scores, sheet 2 has the cluster-aware summary (if --task or "
             "--all-tasks is also provided).",
    )
    parser.add_argument(
        "--manual-clusters",
        metavar="PATH",
        help="Path to a YAML file defining manual dataset clusters. "
             "Produces one table per cluster showing average scores per task.",
    )
    parser.add_argument(
        "--mc-complete-datasets",
        action="store_true",
        help="For manual clusters: within each (cluster, task) group, drop "
             "datasets that not all models have results for.",
    )
    return parser.parse_args()


def main() -> None:
    """Entry point for benchmark clustering analysis."""
    args = parse_args()

    if not args.task and not args.all_tasks and not args.export_excel and not args.manual_clusters:
        print("Error: specify --task <name>, --all-tasks, --export-excel, or --manual-clusters.")
        sys.exit(1)

    task_scores: Dict[str, pd.Series] = {}
    effective_metrics: Dict[str, str] = {}

    if args.task or args.all_tasks:
        tasks = list(TASK_METRICS.keys()) if args.all_tasks else [args.task]

        for task in tasks:
            metric = args.metric if not args.all_tasks else TASK_METRICS[task]
            if metric is None:
                metric = TASK_METRICS[task]
            effective_metrics[task] = metric

            result = analyze_task(
                args.results_dir,
                task,
                metric,
                args.episode_params,
                args.output_dir,
                args.min_models,
                args.include_excluded,
                args.threshold,
                args.complete_datasets,
            )
            if result is not None:
                task_scores[task] = result

        print_summary_table(task_scores, effective_metrics, args.output_dir)
        print(f"Done. Produced analysis for {len(task_scores)}/{len(tasks)} tasks.")

    manual_cluster_tables: Optional[Dict[str, pd.DataFrame]] = None
    manual_cluster_datasets: Optional[Dict[str, Dict[str, List[str]]]] = None
    if args.manual_clusters:
        try:
            clusters = load_manual_clusters(args.manual_clusters)
        except ValueError as e:
            print(f"error loading {args.manual_clusters}: {e}", file=sys.stderr)
            sys.exit(2)
        manual_cluster_tables, manual_cluster_datasets = build_manual_cluster_tables(
            args.results_dir,
            clusters,
            args.episode_params,
            args.include_excluded,
            args.mc_complete_datasets,
        )
        print_manual_cluster_tables(manual_cluster_tables, manual_cluster_datasets, args.output_dir)

    if args.export_excel:
        export_excel(
            args.results_dir,
            args.export_excel,
            args.include_excluded,
            task_scores if task_scores else None,
            effective_metrics if effective_metrics else None,
            manual_cluster_tables,
            manual_cluster_datasets,
        )


if __name__ == "__main__":
    main()
