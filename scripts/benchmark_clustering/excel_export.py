"""Export benchmark scores to a multi-sheet Excel workbook."""
from typing import Dict, List, Optional, Tuple

import pandas as pd

from .config import AGGREGATE_SHEETS, OA_VARIANT_METRICS, TASK_METRICS
from .discovery import _warn_missing_metric, discover_all_scores


def _highlight_best_cells(
    ws,
    cells: List[Tuple[float, int]],
    bold_font,
    underline_font,
    axis: str,
    fixed_idx: int,
) -> None:
    """Bold best and underline second-best cells, handling ties.

    Args:
        ws: The openpyxl worksheet.
        cells: List of (value, index) pairs for the varying axis.
        bold_font: Font to apply to best value(s).
        underline_font: Font to apply to second-best value(s).
        axis: "row" if fixed_idx is a row (varying columns) or
              "col" if fixed_idx is a column (varying rows).
        fixed_idx: The fixed row or column index (1-indexed).
    """
    if len(cells) < 2:
        return
    cells.sort(key=lambda x: x[0], reverse=True)
    best_val = cells[0][0]
    second_val = next((v for v, _ in cells if v < best_val), None)

    for val, idx in cells:
        if val == best_val:
            cell = ws.cell(
                row=fixed_idx if axis == "row" else idx,
                column=idx if axis == "row" else fixed_idx,
            )
            cell.font = bold_font
        elif second_val is not None and val == second_val:
            cell = ws.cell(
                row=fixed_idx if axis == "row" else idx,
                column=idx if axis == "row" else fixed_idx,
            )
            cell.font = underline_font


def export_excel(
    results_dir: str,
    output_path: str,
    include_excluded: bool = False,
    task_scores: Optional[Dict[str, pd.Series]] = None,
    task_metrics: Optional[Dict[str, str]] = None,
    task_n_datasets: Optional[Dict[str, int]] = None,
    manual_cluster_tables: Optional[Dict[str, pd.DataFrame]] = None,
    manual_cluster_datasets: Optional[Dict[str, Dict[str, List[str]]]] = None,
    ranking_plot_paths: Optional[List[str]] = None,
    allowed_models: Optional[set] = None,
) -> None:
    """Export all scores to an Excel file.

    Sheets:
      - "scores": one row per (dataset, task, episode_config) with per-model
        metric values. Best per row is bold, second best is underlined.
      - "summary": cluster-aware aggregated scores per model per task
        (if --task or --all-tasks was used). Includes a "# datasets" row.
      - "ranking": embedded model ranking plot (if ranking_plot_path provided).
      - One sheet per cluster named "mc_{cluster}" with manual cluster
        averages (if --manual-clusters was used). Best per column is bold,
        second best is underlined. Includes dataset list per column.

    Args:
        results_dir: Path to the root results directory.
        output_path: Path for the output .xlsx file.
        include_excluded: If True, include semantic and non-English datasets.
        task_scores: Mapping from task name to per-model aggregated scores.
            If provided, written as the "summary" sheet.
        task_metrics: Mapping from task name to metric name (for column headers).
        task_n_datasets: Mapping from task name to number of datasets used.
        manual_cluster_tables: Mapping from cluster name to DataFrame of manual
            cluster averages (models x tasks).
        manual_cluster_datasets: Mapping from cluster name to dict of column
            name to list of dataset names included in that column.
        ranking_plot_paths: Paths to ranking plot PNGs to embed in Excel
            [ranking.png, ranking_grouped.png].
        allowed_models: If provided, only include models in this set.
    """
    rows = discover_all_scores(results_dir, include_excluded, allowed_models)
    if not rows:
        print("No scores found. Nothing to export.")
        return

    records = []
    warned_missing: set = set()
    for (dataset, task, episode_config), model_metrics in sorted(rows.items()):
        # For order_alignment, emit one row per recognised variant metric so
        # readers can see both acc_mean and distractor_acc_mean side by side.
        # For all other tasks, emit a single row with TASK_METRICS' default.
        if task == "order_alignment":
            metric_keys = sorted(set(OA_VARIANT_METRICS.values()))
        else:
            metric_keys = [TASK_METRICS[task]]

        for primary_metric in metric_keys:
            record: Dict[str, object] = {
                "dataset": dataset,
                "task": task,
                "episode_config": episode_config,
                "primary_metric": primary_metric,
            }
            for model, metrics in model_metrics.items():
                value = metrics.get(primary_metric)
                if value is None:
                    _warn_missing_metric(dataset, task, primary_metric, warned_missing)
                    continue
                record[model] = value
            records.append(record)

    scores_df = pd.DataFrame(records)

    # Ensure metadata columns come first, then models sorted alphabetically
    meta_cols = ["dataset", "task", "episode_config", "primary_metric"]
    model_cols = sorted(c for c in scores_df.columns if c not in meta_cols)
    scores_df = scores_df[meta_cols + model_cols]

    from openpyxl.styles import Font

    BASE_FONT_SIZE = 14
    base_font = Font(size=BASE_FONT_SIZE)
    bold_font = Font(bold=True, size=BASE_FONT_SIZE)
    underline_font = Font(underline="single", size=BASE_FONT_SIZE)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        scores_df[model_cols] = scores_df[model_cols].round(4)
        scores_df.to_excel(writer, sheet_name="scores", index=False)

        # Bold best, underline second best per row (across model columns)
        ws_scores = writer.sheets["scores"]
        model_col_start = len(meta_cols) + 1  # 1-indexed, after meta columns
        model_col_end = model_col_start + len(model_cols) - 1
        for row_idx in range(2, len(scores_df) + 2):  # skip header
            vals = []
            for col_idx in range(model_col_start, model_col_end + 1):
                cell = ws_scores.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    vals.append((cell.value, col_idx))
            _highlight_best_cells(ws_scores, vals, bold_font, underline_font, "row", row_idx)

        if task_scores and task_metrics:
            col_names = {
                task: f"{task} ({task_metrics[task]})"
                for task in task_scores
            }
            columns = {
                col_names[task]: scores
                for task, scores in task_scores.items()
            }
            summary_df = pd.DataFrame(columns)
            summary_df.index.name = "model"
            summary_df = summary_df.round(4)

            # Prepend a "# datasets" row
            ds_counts = task_n_datasets or {}
            n_datasets_row = {
                col_names[task]: ds_counts.get(task, "")
                for task in task_scores
            }
            n_datasets_df = pd.DataFrame(n_datasets_row, index=["# datasets"])
            n_datasets_df.index.name = "model"
            summary_df = pd.concat([n_datasets_df, summary_df])

            summary_df.to_excel(writer, sheet_name="summary")

            # Bold best, underline second best per column (skip # datasets row)
            ws_summary = writer.sheets["summary"]
            # Data starts at row 3 (row 1 = header, row 2 = # datasets)
            for col_idx in range(2, len(summary_df.columns) + 2):  # skip index col
                vals = []
                for row_idx in range(3, len(summary_df) + 2):  # skip header + datasets row
                    cell = ws_summary.cell(row=row_idx, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        vals.append((cell.value, row_idx))
                _highlight_best_cells(ws_summary, vals, bold_font, underline_font, "col", col_idx)

        # Embed ranking plots as sheets
        if ranking_plot_paths:
            from openpyxl.drawing.image import Image as XlImage
            sheet_names = ["ranking", "ranking_grouped"]
            for plot_path, sname in zip(ranking_plot_paths, sheet_names):
                ws = writer.book.create_sheet(sname)
                img = XlImage(plot_path)
                ws.add_image(img, "A1")

        # Aggregate sheets: combine columns from multiple manual cluster tables
        if manual_cluster_tables and AGGREGATE_SHEETS:
            italic_font = Font(italic=True, size=BASE_FONT_SIZE)
            for agg_sheet_name, source_clusters in AGGREGATE_SHEETS:
                avg_series = {}
                for cluster_name in source_clusters:
                    mc_df = manual_cluster_tables.get(cluster_name)
                    if mc_df is None:
                        print(f"  Warning: aggregate sheet '{agg_sheet_name}' references "
                              f"missing cluster '{cluster_name}', skipping it.")
                        continue
                    avg_series[cluster_name] = mc_df.mean(axis=1)
                if not avg_series:
                    continue
                agg_df = pd.DataFrame(avg_series)
                agg_df["average"] = agg_df.mean(axis=1)
                agg_df = agg_df.round(4)
                sheet_name = agg_sheet_name[:31]
                agg_df.to_excel(writer, sheet_name=sheet_name)

                ws = writer.sheets[sheet_name]
                # Bold best, underline second best per column
                for col_idx in range(2, len(agg_df.columns) + 2):
                    vals = []
                    for row_idx in range(2, len(agg_df) + 2):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if isinstance(cell.value, (int, float)):
                            vals.append((cell.value, row_idx))
                    _highlight_best_cells(ws, vals, bold_font, underline_font, "col", col_idx)

        if manual_cluster_tables:
            col_ds_all = manual_cluster_datasets or {}
            for cluster_name, mc_df in sorted(manual_cluster_tables.items()):
                sheet_name = f"mc_{cluster_name}"[:31]  # Excel 31-char limit
                col_ds = col_ds_all.get(cluster_name, {})

                # Find max number of datasets across columns for row offset
                max_datasets = max(
                    (len(ds) for ds in col_ds.values()),
                    default=0,
                )
                # Leave rows for: "Datasets:" label + one row per dataset + blank row
                data_start_row = max_datasets + 2 if max_datasets > 0 else 0
                mc_df = mc_df.copy()
                mc_df["average"] = mc_df.mean(axis=1)
                mc_df = mc_df.round(4)
                mc_df.to_excel(writer, sheet_name=sheet_name, startrow=data_start_row)

                ws = writer.sheets[sheet_name]

                # Write dataset lists above the data
                if col_ds:
                    italic_font = Font(italic=True, size=BASE_FONT_SIZE)
                    ws.cell(row=1, column=1, value="Datasets:").font = italic_font
                    for col_idx, col_name in enumerate(mc_df.columns, start=2):
                        datasets = col_ds.get(col_name, [])
                        for ds_idx, ds_name in enumerate(datasets):
                            cell = ws.cell(row=1 + ds_idx, column=col_idx, value=ds_name)
                            cell.font = italic_font

                # Bold best, underline second best per column
                header_row = data_start_row + 1
                for col_idx in range(2, len(mc_df.columns) + 2):
                    vals = []
                    for row_idx in range(header_row + 1, header_row + len(mc_df) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if isinstance(cell.value, (int, float)):
                            vals.append((cell.value, row_idx))
                    _highlight_best_cells(ws, vals, bold_font, underline_font, "col", col_idx)

        # Set base font size and auto-resize columns for all sheets
        for ws in writer.book.worksheets:
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.font.size is None or cell.font.size < BASE_FONT_SIZE:
                        cell.font = Font(
                            size=BASE_FONT_SIZE,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            underline=cell.font.underline,
                        )
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_len + 2

    n_sheets = 1
    if task_scores:
        n_sheets += 1
    if ranking_plot_paths:
        n_sheets += len(ranking_plot_paths)
    if manual_cluster_tables:
        n_sheets += len(manual_cluster_tables)
    if manual_cluster_tables and AGGREGATE_SHEETS:
        n_sheets += sum(
            1 for _, sources in AGGREGATE_SHEETS
            if any(s in manual_cluster_tables for s in sources)
        )
    print(f"Exported {len(scores_df)} rows × {len(model_cols)} models ({n_sheets} sheets) to {output_path}")
